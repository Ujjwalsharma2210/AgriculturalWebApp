from openpyxl import load_workbook

file = "database.xlsx"

def getInfoExcel(plant_name):

    workbook = load_workbook(filename=file)
    sheet = workbook.active

    plantInfo = []

    for i in range(1, 101):
        a = sheet["B" + str(i)].value
        b = plant_name
        if a == None or b == None:
            a = b = "X"
        elif a.lower() == b.lower():
            plantInfo = []
            for j in range(66, 73):
                plantInfo.append(str(sheet[chr(j) + "1"].value) + " : " + str(sheet[chr(j) + str(i)].value))
                print(plantInfo)
            return plantInfo
    return ""

def getImageURLFromExcel(plant_name):

    workbook = load_workbook(filename=file)
    sheet = workbook.active

    for i in range(1, 110):
        a = sheet["B" + str(i)].value
        b = plant_name
        if a == None or b == None:
            a = b = "X"
        elif a.lower() == b.lower():
            imageURL = str(sheet["O" + str(i)].value)
            return imageURL
    return ""

def getInfoSQL(plant_name):
    pass
